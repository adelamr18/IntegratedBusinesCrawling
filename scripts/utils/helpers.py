from selenium import webdriver
from selenium.webdriver.firefox.service import Service
from selenium.webdriver.firefox.options import Options
import os
from openpyxl import Workbook, load_workbook
from openpyxl import Workbook, load_workbook
from webdriver_manager.firefox import GeckoDriverManager
import csv
import os
# Initialize the Firefox driver
def driver_initialize():
    firefox_options = Options()
    firefox_options.headless = True
    firefox_options.binary_location = '/Applications/Firefox.app/Contents/MacOS/firefox'
    # firefox_options.binary_location = r"C:\Program Files\Mozilla Firefox\firefox.exe"
    service = Service(GeckoDriverManager().install())
    #driver = webdriver.Firefox(service=service, options=firefox_options)
    # service = Service(executable_path=r'C:\\Users\\DiscoCrawler1\\Desktop\\IntegratedBusinesCrawling\\geckodriver.exe')
    driver = webdriver.Firefox(service=service, options=firefox_options)
    return driver


def convert_url_to_arabic(url):
    return url.replace('/en/', '/ar/')

def write_product_to_excel(output_file_name, product):
    try:
        # Check if the output file already exists
        file_exists = os.path.isfile(output_file_name)

        # Load or create a workbook
        if file_exists:
            workbook = load_workbook(output_file_name)
            sheet = workbook.active
        else:
            workbook = Workbook()
            sheet = workbook.active
            # Add headers if creating a new file
            sheet.append([
                'Merchant', 'Id', 'Brand ar', 'Brand en', 'Barcode', 'Item Name AR', 
                'Item Name EN', 'Category 1 EN', 'Category 2 EN', 'Category 3 EN', 
                'Category 4 EN', 'Category 5 EN', 'Category 6 EN', 'Category 7 EN',
                'Category 8 EN', 'Category 9 EN',  
                'Category 1 AR', 'Category 2 AR', 'Category 3 AR', 
                'Category 4 AR', 'Category 5 AR', 'Category 6 AR', 'Category 7 AR',
                'Category 8 AR', 'Category 9 AR',  
                'Price before', 'Price after', 'Offer start date', 'Offer end date', 
                'Url', 'Brand Url', 'Picture', 'Type', 'Crawled on'
            ])

        # Ensure all fields are strings and replace None with empty strings
        product_data = [
            str(getattr(product, field, "")) if getattr(product, field, "") is not None else ""
            for field in [
                'merchant', 'product_id', 'brand_ar', 'brand_en', 'barcode', 
                'name_ar', 'name_en', 'category_one_eng', 'category_two_eng', 'category_three_eng',
                'category_four_eng', 'category_five_eng', 'category_six_eng', 'category_seven_eng',
                'category_eight_eng', 'category_nine_eng',
                'category_one_ar', 'category_two_ar', 'category_three_ar', 
                'category_four_ar', 'category_five_ar', 'category_six_ar', 
                'category_seven_ar', 'category_eight_ar', 'category_nine_ar',
                'price_before', 'price_after', 'offer_start_date', 'offer_end_date', 
                'url', 'brand_image_url', 'image_url', 'source_type', 'crawled_on'
            ]
        ]

        # Append the processed product data to the sheet
        sheet.append(product_data)

        # Save and close the workbook
        workbook.save(output_file_name)
        workbook.close()

    except Exception as e:
        print(f"Error writing to Excel file {output_file_name}: {e}")


def write_brands_to_excel(output_file_name, brands):
    try:
        # Check if the output file already exists
        file_exists = os.path.isfile(output_file_name)

        # Load or create a workbook
        if file_exists:
            workbook = load_workbook(output_file_name)
            sheet = workbook.active
        else:
            workbook = Workbook()
            sheet = workbook.active
            # Add headers if creating a new file
            sheet.append(['Brand Name', 'Brand Image URL'])

        # Write each brand's name and image URL to the sheet
        for brand_name, image_url in brands:
            sheet.append([brand_name, image_url])

        # Save and close the workbook
        workbook.save(output_file_name)
        workbook.close()

    except Exception as e:
        print(f"Error writing to Excel file {output_file_name}: {e}")
        
def read_urls_from_csv(csv_file_path):
    urls = []
    try:
        with open(csv_file_path, mode='r', newline='', encoding='utf-8') as file:
            reader = csv.reader(file)
            # Skip the header
            next(reader)
            for row in reader:
                # Append URLs where is_processed is either False, empty, or not present
                if len(row) < 3 or row[2].strip().lower() == 'false' or row[2].strip() == '':
                    urls.append(row[1])  # Append the URL
    except Exception as e:
        print(f"Error reading CSV file: {e}")
    return urls

# Update the processed status of a URL in the CSV
def update_is_processed_in_csv(url, is_successful, input_csv_path):
    rows = []
    url_found = False  # Flag to check if the URL exists
    try:
        with open(input_csv_path, mode='r', newline='', encoding='utf-8') as file:
            reader = csv.reader(file)
            header = next(reader)

            # Add is_processed header if not exists
            if len(header) < 3 or header[2] != 'is_processed':
                header.append('is_processed')
            rows.append(header)

            for row in reader:
                if row[1] == url:
                    url_found = True
                    # If URL is found, update the is_processed status
                    if len(row) < 3:
                        row.append('True' if is_successful else 'False')
                    else:
                        row[2] = 'True' if is_successful else 'False'
                rows.append(row)

            # If the URL was not found, append a new row with the corresponding status
            if not url_found:
                rows.append([None, url, 'True' if is_successful else 'False'])  # Assuming None for the first column

        # Write updated rows back to CSV
        with open(input_csv_path, mode='w', newline='', encoding='utf-8') as file:
            writer = csv.writer(file)
            writer.writerows(rows)

    except Exception as e:
        print(f"Error updating CSV file: {e}")

        # If an error happens, we still want to mark the URL as not processed
        rows.append([None, url, 'False'])  # Assuming None for the first column
        # Write rows back to CSV, ensuring the header is included
        try:
            with open(input_csv_path, mode='w', newline='', encoding='utf-8') as file:
                writer = csv.writer(file)
                writer.writerows(rows)
        except Exception as write_error:
            print(f"Error writing to CSV file: {write_error}")


    except Exception as e:
        print(f"Error updating CSV file: {e}")
        # If an error happens, we still want to mark the URL as not processed
        rows.append([None, url, 'False'])  # Assuming None for the first column
        # Write rows back to CSV, ensuring the header is included
        try:
            with open(input_csv_path, mode='w', newline='', encoding='utf-8') as file:
                writer = csv.writer(file)
                writer.writerows(rows)
        except Exception as write_error:
            print(f"Error writing to CSV file: {write_error}")        
