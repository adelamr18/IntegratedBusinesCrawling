import re
import os
from datetime import datetime
from selenium.webdriver.common.by import By
from utils.helpers import driver_intialize, convert_url_to_arabic
from openpyxl import Workbook, load_workbook
import csv
import re
from datetime import datetime, timedelta

def extract_offer_end_date(driver):
    try:
        # Find the second child element that contains the number of days
        element = driver.find_element(By.CSS_SELECTOR, '.css-juexlj > span:nth-child(2)')
        
        # Extract the number of days from the text (e.g., "2 days")
        days_text = element.text.strip()
        
        # Use regex to find the number in the text
        match = re.search(r'\d+', days_text)
        
        if match:
            # Convert the extracted number to an integer
            days_to_add = int(match.group(0))
            
            # Calculate the new date by adding the number of days to today's date
            calculated_date = datetime.now() + timedelta(days=days_to_add)
            
            # Return the calculated date in the format YYYY-MM-DD
            return calculated_date.strftime('%Y-%m-%d')
        
        return ""  # Return empty string if no match found

    except Exception as e:
        return ""
    
def extract_parent_category(driver):
    try:

        elements = driver.find_elements(By.CSS_SELECTOR, '.css-iamwo8')
        
        # Skip the first and last elements
        selected_elements = elements[1:-1]

        # Read text from each element and handle empty texts
        parent_categories = []
        for element in selected_elements:
            text = element.text.strip()
            # Append the text or an empty string if the text is empty
            parent_categories.append(text if text else "")

        # Join the results with '-'
        formatted_result = ' - '.join(parent_categories)
        
        return formatted_result if formatted_result else "Parent categories not found"

    except Exception as e:
        print(f"Error extracting parent categories: {e}")
        return "Parent categories not found"

def extract_child_category(driver):
    try:
        child_category = driver.find_element(By.CSS_SELECTOR, '.css-82hvi6').text
        
        if not child_category:
            return "Child category not found"

        return child_category

    except Exception as e:
        print(f"Error extracting child category: {e}")
        return "Child category not found"
    
def extract_product_barcode(driver):
    try:
         
        element = driver.find_element(By.CSS_SELECTOR, "#__next > div.css-qo9h12 > main > div > div.css-9p8u88 > div:nth-child(2) > script")
        barcode = element.get_attribute("data-flix-ean")
    
        return barcode if barcode else "product barcode not found"
    
    except Exception:
        return "product barcode not found"

def extract_product_id(url):
    driver.get(url)
    match = re.search(r'/p/(\d+)', url)
    
    return match.group(1) if match else "id not found"

def extract_product_name_in_arabic(driver, url):
    driver.get(url)
    
    try:
        product_name_ar = driver.find_element(By.CSS_SELECTOR, '.css-106scfp').text
        
        if not product_name_ar:
            return "لم يتم العثور على اسم المنتج"

        return product_name_ar
    

    except Exception as e:
        print(f"Error extracting product name: {e}")
        return "لم يتم العثور على اسم المنتج"
def extract_image_url(driver):
    
    try:
        # Select the div with the specified class
        image_div = driver.find_element(By.CSS_SELECTOR, 'div.css-1c2pck7 img')
        
        # Get the src attribute of the img tag
        img_url = image_div.get_attribute('src')
        return img_url

    except Exception as e:
        print(f"Error extracting image URL: {e}")
        return "Image not found"        
        
def extract_product_name_in_english(driver):
    try:
        product_name_ar = driver.find_element(By.CSS_SELECTOR, '.css-106scfp').text
        
        if not product_name_ar:
            return "Product name not found"

        return product_name_ar

    except Exception as e:
        print(f"Error extracting product name: {e}")
        return "Product name not found"
      
def extract_product_price_before_offer(driver):    
    price_text = ""

    try:
        # Try to find the price element with the offer
        price_element_with_offer = driver.find_element(By.CSS_SELECTOR, '.css-1jh6byp')
        price_text = price_element_with_offer.text
        
        if price_text:
            print('Offer price found:', price_text)
            price_element_with_offer_text = driver.find_element(By.CSS_SELECTOR, 'del.css-1bdwabt').text
            
            if 'Use code' in price_element_with_offer_text:
                raise Exception("Promotional code found, exiting...")    
            
            print('Price before offer:', price_element_with_offer_text)
            
            # Extract only the numeric part from the price before the offer
            match = re.search(r'\d+\.\d+', price_element_with_offer_text)
            print(match.group(0))
            return match.group(0) if match else "Price not found"

    except Exception as e:
        print("Offer price element not found or promotional code detected, trying to get regular price...")
        try:
            price_element = driver.find_element(By.CSS_SELECTOR, '.css-17ctnp')
            price_text = price_element.text
            
            # Extract the numeric part
            match = re.search(r'\d+\.\d+', price_text)
            return match.group(0) if match else "Price not found"

        except Exception:
            try:
                price_element = driver.find_element(By.XPATH, "//h2[contains(text(), 'EGP')]")
                price_text = price_element.text
                
                # Extract the numeric part
                match = re.search(r'\d+\.\d+', price_text)
                return match.group(0) if match else "Price not found"
                
            except Exception:
                return "Price not found"

    return "Price not found"

def extract_product_price_after_offer(driver):
    price_text = ""
    try:
        price_element = driver.find_element(By.CSS_SELECTOR, '.css-1i90gmp')
        price_text = price_element.text
        
        # Extract the numeric part
        match = re.search(r'\d+\.\d+', price_text)
        return match.group(0) if match else ""

    except Exception:
        return ""

def write_to_excel(output_file_name, product_id, product_barcode, child_category, parent_category, product_name_in_arabic, product_name_in_english, price_before_offer, price_after_offer, offer_end_data, image_url, url):
    # Check if the file exists
    file_exists = os.path.isfile(output_file_name)
    
    does_offer_exist = price_after_offer or False
    offer_start_date = datetime.now().strftime('%Y-%m-%d') if does_offer_exist else ''

    if file_exists:
        workbook = load_workbook(output_file_name)
        sheet = workbook.active
    else:
        workbook = Workbook()
        sheet = workbook.active
        
        # Write the header if the file doesn't exist
        sheet.append([
            'Merchant', 'Id', 'Brand ar', 'Brand en', 'Barcode', 'Item Name AR', 
            'Item Name EN', 'Category', 'Parent Category', 'Price before', 
            'Price after', 'Offer start date', 'Offer end date', 'Url', 
            'Picture', 'Type', 'Crawled on'
        ])

    # Writing data into Excel
    sheet.append([
        'Carrefour',             # Merchant
        product_id,              # Id
        '',                      # Brand ar 
        '',                      # Brand en
        product_barcode,         # Barcode
        product_name_in_arabic,  # Item Name AR
        product_name_in_english, # Item Name EN
        child_category,          # Category
        parent_category,         # Parent Category
        price_before_offer,      # Price before offer
        price_after_offer,       # Price after offer
        offer_start_date,        # Offer start date
        offer_end_data,          # Offer end date
        url,                     # Product url
        image_url,               # Product picture
        'Website',               # Type of information source
        datetime.now().strftime('%Y-%m-%d %H:%M:%S')  # crawled on timestamp
    ])

    # Save the workbook after each append
    workbook.save(output_file_name)

def process_urls_and_save_to_excel(csv_file, output_file, driver):
    
    todays_date = datetime.now().strftime('%d_%m_%Y')
    output_file_name = os.path.join(output_file, f"extract_carrefour_data_{todays_date}.xlsx")

    try:
        with open(csv_file, mode='r') as file:
            reader = csv.DictReader(file)
            for row in reader:
                url = row['URL']
                formatted_url_in_arabic = convert_url_to_arabic(url)
                
                # Extract product name in Arabic
                driver.get(formatted_url_in_arabic)
                product_name_in_arabic = extract_product_name_in_arabic(driver, formatted_url_in_arabic)
            
                # Extract product name in English
                driver.get(url)
                product_name_in_english = extract_product_name_in_english(driver)
                
                # Extract product id
                product_id = extract_product_id(url)
                
                # Extract child category of product
                child_category = extract_child_category(driver)
                
                # Extract parent category of product
                parent_category = extract_parent_category(driver)
                
                # Extract product barcode
                product_barcode = extract_product_barcode(driver)
                
                # Extract the price for each URL before offer
                price_before_offer = extract_product_price_before_offer(driver)
                
                # Extract the price for each url after offer if exists
                price_after_offer = extract_product_price_after_offer(driver)
                
                # Extract offer end date
                offer_end_data = extract_offer_end_date(driver)
                
                # Extract image url
                image_url = extract_image_url(driver)
                
                # Write to Excel directly for each URL
                write_to_excel(output_file_name, product_id, product_barcode, child_category, parent_category, product_name_in_arabic, product_name_in_english, price_before_offer, price_after_offer, offer_end_data, image_url, url)
                

        print(f"Data successfully saved to {output_file_name}")
        driver.quit()
    except Exception as e:
        print(f"An error occurred during processing: {e}")

# Call the function with the appropriate CSV file path and output directory
driver = driver_intialize()
input_csv_path = r'../extractions/extract_carrefour_urls_19_09_2024.csv'
output_directory = r'../extractions'
process_urls_and_save_to_excel(input_csv_path, output_directory, driver)

